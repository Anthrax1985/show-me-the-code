
import xlrd
import xlwt
import openpyxl
import json

def write_03_student(spath, dic):
    book = xlwt.Workbook()
    sheet = book.add_sheet('student')
    idx_r = 0
    for r in dic:
        idx_c = 0
        sheet.write(idx_r, idx_c, str(r))
        for c in range(1, len(dic[r]) + 1):
            sheet.write(idx_r, idx_c + c, str(dic[r][c - 1]))
        idx_r += 1
    book.save(spath)
    print('03_exce_done')


def write_07_student(spath, dic):
    book = openpyxl.Workbook()
    book.remove(book.get_active_sheet())
    sheet = book.create_sheet('student', 0)
    idx_r = 1
    for r in dic:
        idx_c = 1
        sheet.cell(idx_r, idx_c, str(r))
        for c in range(1, len(dic[r]) + 1):
            sheet.cell(idx_r, idx_c + c, str(dic[r][c - 1]))
        idx_r += 1
    book.save(spath)
    print('07_exce_done')


def get_txt_content():
    with open('student.txt', 'r') as f:
        return f.read().encode('utf-8')


if __name__ == '__main__':
    text = get_txt_content()
    dic = json.loads(text)
    write_03_student('03book.xls', dic)
    write_07_student('07book.xlsx', dic)